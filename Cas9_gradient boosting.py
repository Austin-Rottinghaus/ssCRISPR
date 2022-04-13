# -*- coding: utf-8 -*-
"""
Created on Mon Jul 20 20:49:00 2020

@author: Austin Rottinghaus
"""

from seqfold import *
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from Bio.SeqUtils import MeltingTemp as mt
from Bio.Seq import Seq
import sklearn
import pickle

import matplotlib.pyplot as plt
from sklearn import datasets, ensemble
from sklearn.inspection import permutation_importance
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split

# Extract scores and sgRNA property data from excel sheet
wb=load_workbook('sgRNAs_prediction.xlsx')
ws=wb['sgRNA_scores']

#wb_results=load_workbook('sgRNAs_prediction_expanded_results2.xlsx')
#ws_predict=wb_results['Prediction']
#ws_importance=wb_results['Perm_Importance']

log2scores=[] # the y variable
sgRNA_properties=[] # the x variable
property_names=[]

for i in range(6,402):
    property_names.append(ws.cell(1,i).value)

for i in range(2,56337):
    log2scores.append(ws.cell(i,2).value)
    new_properties=[]
    for j in range(6,402):
        new_properties.append(ws.cell(i,j).value)
    sgRNA_properties.append(new_properties)

# Split data into 90% train and 10% test    
sgRNA_properties_train, sgRNA_properties_test, log2scores_train, log2scores_test = train_test_split(
    sgRNA_properties, log2scores, test_size=0.1, random_state=13)



# Set parameters for the gradient boosting regression
params = {'n_estimators': 750, #
          'max_depth': 8, #
          'min_samples_split': 75, #
          'learning_rate': .03, #
          'loss': 'ls'}

# Perform the regression
reg = ensemble.GradientBoostingRegressor(**params)
reg.fit(sgRNA_properties_train, log2scores_train)
print(750)
mse = mean_squared_error(log2scores_test, reg.predict(sgRNA_properties_test))
print("The mean squared error (MSE) on test set: {:.4f}".format(mse))


pickle.dump(reg, open('finalized_model.sav', 'wb'))

#all_predictions = reg.predict(sgRNA_properties)
#for i in range(2,56337):
#    ws_predict.cell(i,3).value=all_predictions[i-2]


# Plot the deviance throughout the successive boosting iterations
test_score = np.zeros((params['n_estimators'],), dtype=np.float64)
for i, log2scores_pred in enumerate(reg.staged_predict(sgRNA_properties_test)):
    test_score[i] = reg.loss_(log2scores_test, log2scores_pred)

fig = plt.figure(figsize=(6, 6))
plt.subplot(1, 1, 1)
plt.title('Deviance')
plt.plot(np.arange(params['n_estimators']) + 1, reg.train_score_, 'b-',
         label='Training Set Deviance')
plt.plot(np.arange(params['n_estimators']) + 1, test_score, 'r-',
         label='Test Set Deviance')
plt.legend(loc='upper right')
plt.xlabel('Boosting Iterations')
plt.ylabel('Deviance')
fig.tight_layout()
plt.show()



# Plot feature importance
#feature_importance = reg.feature_importances_
#sorted_idx = np.argsort(feature_importance)
#pos = np.arange(sorted_idx.shape[0]) + .5
#fig = plt.figure(figsize=(12, 6))
#plt.subplot(1, 2, 1)
#plt.barh(pos, feature_importance[sorted_idx], align='center')
#plt.yticks(pos, np.array(property_names))
#plt.title('Feature Importance (MDI)')

#result = permutation_importance(reg, sgRNA_properties_test, log2scores_test, n_repeats=10,
#                                random_state=42, n_jobs=2)
#sorted_idx = result.importances_mean.argsort()
#plt.subplot(1, 2, 2)
#plt.boxplot(result.importances[sorted_idx].T,
#            vert=False, labels=np.array(property_names)[sorted_idx])
#plt.title("Permutation Importance (test set)")
#fig.tight_layout()
#plt.show()


wb.close()
#wb_results.save('sgRNAs_prediction_expanded_results2.xlsx')