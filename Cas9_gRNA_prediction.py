# -*- coding: utf-8 -*-
"""
Created on Wed Aug 26 15:40:08 2020

@author: austr
"""
from openpyxl import load_workbook
import numpy
import pickle

wb=load_workbook('sgRNAs_vibrio.xlsx')
ws=wb['Data']

all_gRNA_properties=[]
for i in range(2,1363): #56337
    gRNA_properties=[]
    for j in range(6,402):
        gRNA_properties.append(ws.cell(i,j).value)
    all_gRNA_properties.append(gRNA_properties)

prediction_model = pickle.load(open('finalized_model.sav','rb'))
gRNA_efficiencies=prediction_model.predict(all_gRNA_properties)

for i in range(len(gRNA_efficiencies)):
    ws.cell(i+2,405).value=gRNA_efficiencies[i]

wb.save('sgRNAs_vibrio.xlsx')