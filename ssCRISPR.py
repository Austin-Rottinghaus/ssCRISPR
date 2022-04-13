# Code by Austin Rottinghaus
############################################################

from Bio import SeqIO, Entrez
from Bio.SeqUtils import nt_search
from Bio.Seq import Seq
from Bio.Alphabet import generic_dna, DNAAlphabet
import re
import sys
import os
from openpyxl import load_workbook, cell, worksheet, workbook
import numpy
import pickle
import random
from Bio.SeqUtils import MeltingTemp as mt
from seqfold import *
from itertools import compress 
import math
import time
import sklearn
from sklearn.ensemble import GradientBoostingRegressor

# determine if the application is a frozen `.exe` (e.g. pyinstaller --onefile) 
if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)
# or a script file (e.g. `.py` / `.pyw`)
elif __file__:
    application_path = os.path.dirname(__file__)

sleep_time=3
max_gRNAs_2bp=100000
max_gRNAs_3bp=5000
max_gRNAs=5000 #For thermodynamic efficiency predictions
max_gRNAs_4bp=100

excel_name = 'All_NCBI_Strains.xlsx'
excel_path = os.path.join(application_path, excel_name)
wb_strains=load_workbook(excel_path)
ws_strains=wb_strains['All NCBI Strains']

def create_strains_list(desired_letter):
    all_strains_NCBInumbers=[]
    for i in range(2,27570): #19027
        if str(ws_strains.cell(i,1).value)[0]==desired_letter:
            new_org_name=str(ws_strains.cell(i,1).value)
            new_str_name=str(ws_strains.cell(i,2).value)
            if new_str_name not in new_org_name:
                new_org_name=new_org_name+' '+new_str_name

            if any(new_org_name in sublist for sublist in all_strains_NCBInumbers) is False:
                new_str_replicons=str(ws_strains.cell(i,4).value)

                number_chrom=new_str_replicons.count('chromosome')

                if number_chrom == 1:
                    location_chrom_start=new_str_replicons.index(':')+1
                    location_chrom_end=new_str_replicons.index('.')+2
                    new_str_replicons=new_str_replicons[location_chrom_start:location_chrom_end]
                    all_strains_NCBInumbers.append([str(new_org_name),str(new_str_replicons)])
                else:
                    new_str_replicons_list=[]
                    while number_chrom>0:
                        location_chrom_start=new_str_replicons.index(':')+1
                        new_str_replicons=new_str_replicons[location_chrom_start-2:]
                        location_chrom_start=new_str_replicons.index(':')+1
                        location_chrom_end=new_str_replicons.index('.')+2
                        new_str_replicons_list.append(str(new_str_replicons[location_chrom_start:location_chrom_end]))
                        new_str_replicons=new_str_replicons[location_chrom_end+2:]
                        number_chrom=number_chrom-1
                    all_strains_NCBInumbers.append([new_org_name,new_str_replicons_list])
    return(all_strains_NCBInumbers)

A_strains_numbers=create_strains_list('A')
A_strains_list=[]
for strain in A_strains_numbers:
    A_strains_list.append(strain[0])

B_strains_numbers=create_strains_list('B')
B_strains_list=[]
for strain in B_strains_numbers:
    B_strains_list.append(strain[0])

C_strains_numbers=create_strains_list('C')
C_strains_list=[]
for strain in C_strains_numbers:
    C_strains_list.append(strain[0])

D_strains_numbers=create_strains_list('D')
D_strains_list=[]
for strain in D_strains_numbers:
    D_strains_list.append(strain[0])

E_strains_numbers=create_strains_list('E')
E_strains_list=[]
for strain in E_strains_numbers:
    E_strains_list.append(strain[0])

F_strains_numbers=create_strains_list('F')
F_strains_list=[]
for strain in F_strains_numbers:
    F_strains_list.append(strain[0])

G_strains_numbers=create_strains_list('G')
G_strains_list=[]
for strain in G_strains_numbers:
    G_strains_list.append(strain[0])

H_strains_numbers=create_strains_list('H')
H_strains_list=[]
for strain in H_strains_numbers:
    H_strains_list.append(strain[0])

I_strains_numbers=create_strains_list('I')
I_strains_list=[]
for strain in I_strains_numbers:
    I_strains_list.append(strain[0])

J_strains_numbers=create_strains_list('J')
J_strains_list=[]
for strain in J_strains_numbers:
    J_strains_list.append(strain[0])

K_strains_numbers=create_strains_list('K')
K_strains_list=[]
for strain in K_strains_numbers:
    K_strains_list.append(strain[0])

L_strains_numbers=create_strains_list('L')
L_strains_list=[]
for strain in L_strains_numbers:
    L_strains_list.append(strain[0])

M_strains_numbers=create_strains_list('M')
M_strains_list=[]
for strain in M_strains_numbers:
    M_strains_list.append(strain[0])

N_strains_numbers=create_strains_list('N')
N_strains_list=[]
for strain in N_strains_numbers:
    N_strains_list.append(strain[0])

O_strains_numbers=create_strains_list('O')
O_strains_list=[]
for strain in O_strains_numbers:
    O_strains_list.append(strain[0])

P_strains_numbers=create_strains_list('P')
P_strains_list=[]
for strain in P_strains_numbers:
    P_strains_list.append(strain[0])

Q_strains_numbers=create_strains_list('Q')
Q_strains_list=[]
for strain in Q_strains_numbers:
    Q_strains_list.append(strain[0])

R_strains_numbers=create_strains_list('R')
R_strains_list=[]
for strain in R_strains_numbers:
    R_strains_list.append(strain[0])

S_strains_numbers=create_strains_list('S')
S_strains_list=[]
for strain in S_strains_numbers:
    S_strains_list.append(strain[0])

T_strains_numbers=create_strains_list('T')
T_strains_list=[]
for strain in T_strains_numbers:
    T_strains_list.append(strain[0])

U_strains_numbers=create_strains_list('U')
U_strains_list=[]
for strain in U_strains_numbers:
    U_strains_list.append(strain[0])

V_strains_numbers=create_strains_list('V')
V_strains_list=[]
for strain in V_strains_numbers:
    V_strains_list.append(strain[0])

W_strains_numbers=create_strains_list('W')
W_strains_list=[]
for strain in W_strains_numbers:
    W_strains_list.append(strain[0])

X_strains_numbers=create_strains_list('X')
X_strains_list=[]
for strain in X_strains_numbers:
    X_strains_list.append(strain[0])

Y_strains_numbers=create_strains_list('Y')
Y_strains_list=[]
for strain in Y_strains_numbers:
    Y_strains_list.append(strain[0])

Z_strains_numbers=create_strains_list('Z')
Z_strains_list=[]
for strain in Z_strains_numbers:
    Z_strains_list.append(strain[0])

all_strains_numbers=[A_strains_numbers,B_strains_numbers,C_strains_numbers,D_strains_numbers,E_strains_numbers,F_strains_numbers,G_strains_numbers,H_strains_numbers,
                    I_strains_numbers,J_strains_numbers,K_strains_numbers,L_strains_numbers,M_strains_numbers,N_strains_numbers,O_strains_numbers,P_strains_numbers,
                    Q_strains_numbers,R_strains_numbers,S_strains_numbers,T_strains_numbers,U_strains_numbers,V_strains_numbers,W_strains_numbers,X_strains_numbers,
                    Y_strains_numbers,Z_strains_numbers]
all_strains_numbers_flat=[val for sublist in all_strains_numbers for val in sublist]
all_strains_numbers_flat2=[val for sublist in all_strains_numbers_flat for val in sublist]

# Function to create gRNAs with single bp mismatches to get rid of ones that have less than 2bp mismatches
def replace_nucs(sequence):
    new_sequences=[]
    for i in range(20):
        new_sequences.append(sequence[0:i]+'A'+sequence[i+1:])
        new_sequences.append(sequence[0:i]+'T'+sequence[i+1:])
        new_sequences.append(sequence[0:i]+'G'+sequence[i+1:])
        new_sequences.append(sequence[0:i]+'C'+sequence[i+1:])
    return new_sequences

def replace_nucs2bp(sequence):
    new_sequences_3bp=[]
    for i in range(20):
        new_sequence1=sequence[0:i]+'A'+sequence[i+1:]
        new_sequence2=sequence[0:i]+'T'+sequence[i+1:]
        new_sequence3=sequence[0:i]+'G'+sequence[i+1:]
        new_sequence4=sequence[0:i]+'C'+sequence[i+1:]
        for j in range(20):
            new_sequences_3bp.append(new_sequence1[0:j]+'A'+new_sequence1[j+1:])
            new_sequences_3bp.append(new_sequence1[0:j]+'T'+new_sequence1[j+1:])
            new_sequences_3bp.append(new_sequence1[0:j]+'G'+new_sequence1[j+1:])
            new_sequences_3bp.append(new_sequence1[0:j]+'C'+new_sequence1[j+1:])
            new_sequences_3bp.append(new_sequence2[0:j]+'A'+new_sequence2[j+1:])
            new_sequences_3bp.append(new_sequence2[0:j]+'T'+new_sequence2[j+1:])
            new_sequences_3bp.append(new_sequence2[0:j]+'G'+new_sequence2[j+1:])
            new_sequences_3bp.append(new_sequence2[0:j]+'C'+new_sequence2[j+1:])
            new_sequences_3bp.append(new_sequence3[0:j]+'A'+new_sequence3[j+1:])
            new_sequences_3bp.append(new_sequence3[0:j]+'T'+new_sequence3[j+1:])
            new_sequences_3bp.append(new_sequence3[0:j]+'G'+new_sequence3[j+1:])
            new_sequences_3bp.append(new_sequence3[0:j]+'C'+new_sequence3[j+1:])
            new_sequences_3bp.append(new_sequence4[0:j]+'A'+new_sequence4[j+1:])
            new_sequences_3bp.append(new_sequence4[0:j]+'T'+new_sequence4[j+1:])
            new_sequences_3bp.append(new_sequence4[0:j]+'G'+new_sequence4[j+1:])
            new_sequences_3bp.append(new_sequence4[0:j]+'C'+new_sequence4[j+1:])
    return new_sequences_3bp

def replace_nucs3bp(sequence):
    new_sequences_4bp=[]
    for i in range(20):
        new_sequenceA=sequence[0:i]+'A'+sequence[i+1:]
        new_sequenceT=sequence[0:i]+'T'+sequence[i+1:]
        new_sequenceG=sequence[0:i]+'G'+sequence[i+1:]
        new_sequenceC=sequence[0:i]+'C'+sequence[i+1:]
        for j in range(20):
            new_sequence1=new_sequenceA[0:j]+'A'+new_sequenceA[j+1:]
            new_sequence2=new_sequenceA[0:j]+'T'+new_sequenceA[j+1:]
            new_sequence3=new_sequenceA[0:j]+'G'+new_sequenceA[j+1:]
            new_sequence4=new_sequenceA[0:j]+'C'+new_sequenceA[j+1:]
            new_sequence5=new_sequenceT[0:j]+'A'+new_sequenceT[j+1:]
            new_sequence6=new_sequenceT[0:j]+'T'+new_sequenceT[j+1:]
            new_sequence7=new_sequenceT[0:j]+'G'+new_sequenceT[j+1:]
            new_sequence8=new_sequenceT[0:j]+'C'+new_sequenceT[j+1:]
            new_sequence9=new_sequenceG[0:j]+'A'+new_sequenceG[j+1:]
            new_sequence10=new_sequenceG[0:j]+'T'+new_sequenceG[j+1:]
            new_sequence11=new_sequenceG[0:j]+'G'+new_sequenceG[j+1:]
            new_sequence12=new_sequenceG[0:j]+'C'+new_sequenceG[j+1:]
            new_sequence13=new_sequenceC[0:j]+'A'+new_sequenceC[j+1:]
            new_sequence14=new_sequenceC[0:j]+'T'+new_sequenceC[j+1:]
            new_sequence15=new_sequenceC[0:j]+'G'+new_sequenceC[j+1:]
            new_sequence16=new_sequenceC[0:j]+'C'+new_sequenceC[j+1:]
            for k in range(20):
                new_sequences_4bp.append(new_sequence1[0:k]+'A'+new_sequence1[k+1:])
                new_sequences_4bp.append(new_sequence1[0:k]+'T'+new_sequence1[k+1:])
                new_sequences_4bp.append(new_sequence1[0:k]+'G'+new_sequence1[k+1:])
                new_sequences_4bp.append(new_sequence1[0:k]+'C'+new_sequence1[k+1:])
                new_sequences_4bp.append(new_sequence2[0:k]+'A'+new_sequence2[k+1:])
                new_sequences_4bp.append(new_sequence2[0:k]+'T'+new_sequence2[k+1:])
                new_sequences_4bp.append(new_sequence2[0:k]+'G'+new_sequence2[k+1:])
                new_sequences_4bp.append(new_sequence2[0:k]+'C'+new_sequence2[k+1:])
                new_sequences_4bp.append(new_sequence3[0:k]+'A'+new_sequence3[k+1:])
                new_sequences_4bp.append(new_sequence3[0:k]+'T'+new_sequence3[k+1:])
                new_sequences_4bp.append(new_sequence3[0:k]+'G'+new_sequence3[k+1:])
                new_sequences_4bp.append(new_sequence3[0:k]+'C'+new_sequence3[k+1:])
                new_sequences_4bp.append(new_sequence4[0:k]+'A'+new_sequence4[k+1:])
                new_sequences_4bp.append(new_sequence4[0:k]+'T'+new_sequence4[k+1:])
                new_sequences_4bp.append(new_sequence4[0:k]+'G'+new_sequence4[k+1:])
                new_sequences_4bp.append(new_sequence4[0:k]+'C'+new_sequence4[k+1:])
                new_sequences_4bp.append(new_sequence5[0:k]+'A'+new_sequence5[k+1:])
                new_sequences_4bp.append(new_sequence5[0:k]+'T'+new_sequence5[k+1:])
                new_sequences_4bp.append(new_sequence5[0:k]+'G'+new_sequence5[k+1:])
                new_sequences_4bp.append(new_sequence5[0:k]+'C'+new_sequence5[k+1:])
                new_sequences_4bp.append(new_sequence6[0:k]+'A'+new_sequence6[k+1:])
                new_sequences_4bp.append(new_sequence6[0:k]+'T'+new_sequence6[k+1:])
                new_sequences_4bp.append(new_sequence6[0:k]+'G'+new_sequence6[k+1:])
                new_sequences_4bp.append(new_sequence6[0:k]+'C'+new_sequence6[k+1:])
                new_sequences_4bp.append(new_sequence7[0:k]+'A'+new_sequence7[k+1:])
                new_sequences_4bp.append(new_sequence7[0:k]+'T'+new_sequence7[k+1:])
                new_sequences_4bp.append(new_sequence7[0:k]+'G'+new_sequence7[k+1:])
                new_sequences_4bp.append(new_sequence7[0:k]+'C'+new_sequence7[k+1:])
                new_sequences_4bp.append(new_sequence8[0:k]+'A'+new_sequence8[k+1:])
                new_sequences_4bp.append(new_sequence8[0:k]+'T'+new_sequence8[k+1:])
                new_sequences_4bp.append(new_sequence8[0:k]+'G'+new_sequence8[k+1:])
                new_sequences_4bp.append(new_sequence8[0:k]+'C'+new_sequence8[k+1:])
                new_sequences_4bp.append(new_sequence9[0:k]+'A'+new_sequence9[k+1:])
                new_sequences_4bp.append(new_sequence9[0:k]+'T'+new_sequence9[k+1:])
                new_sequences_4bp.append(new_sequence9[0:k]+'G'+new_sequence9[k+1:])
                new_sequences_4bp.append(new_sequence9[0:k]+'C'+new_sequence9[k+1:])
                new_sequences_4bp.append(new_sequence10[0:k]+'A'+new_sequence10[k+1:])
                new_sequences_4bp.append(new_sequence10[0:k]+'T'+new_sequence10[k+1:])
                new_sequences_4bp.append(new_sequence10[0:k]+'G'+new_sequence10[k+1:])
                new_sequences_4bp.append(new_sequence10[0:k]+'C'+new_sequence10[k+1:])
                new_sequences_4bp.append(new_sequence11[0:k]+'A'+new_sequence11[k+1:])
                new_sequences_4bp.append(new_sequence11[0:k]+'T'+new_sequence11[k+1:])
                new_sequences_4bp.append(new_sequence11[0:k]+'G'+new_sequence11[k+1:])
                new_sequences_4bp.append(new_sequence11[0:k]+'C'+new_sequence11[k+1:])
                new_sequences_4bp.append(new_sequence12[0:k]+'A'+new_sequence12[k+1:])
                new_sequences_4bp.append(new_sequence12[0:k]+'T'+new_sequence12[k+1:])
                new_sequences_4bp.append(new_sequence12[0:k]+'G'+new_sequence12[k+1:])
                new_sequences_4bp.append(new_sequence12[0:k]+'C'+new_sequence12[k+1:])
                new_sequences_4bp.append(new_sequence13[0:k]+'A'+new_sequence13[k+1:])
                new_sequences_4bp.append(new_sequence13[0:k]+'T'+new_sequence13[k+1:])
                new_sequences_4bp.append(new_sequence13[0:k]+'G'+new_sequence13[k+1:])
                new_sequences_4bp.append(new_sequence13[0:k]+'C'+new_sequence13[k+1:])
                new_sequences_4bp.append(new_sequence14[0:k]+'A'+new_sequence14[k+1:])
                new_sequences_4bp.append(new_sequence14[0:k]+'T'+new_sequence14[k+1:])
                new_sequences_4bp.append(new_sequence14[0:k]+'G'+new_sequence14[k+1:])
                new_sequences_4bp.append(new_sequence14[0:k]+'C'+new_sequence14[k+1:])
                new_sequences_4bp.append(new_sequence15[0:k]+'A'+new_sequence15[k+1:])
                new_sequences_4bp.append(new_sequence15[0:k]+'T'+new_sequence15[k+1:])
                new_sequences_4bp.append(new_sequence15[0:k]+'G'+new_sequence15[k+1:])
                new_sequences_4bp.append(new_sequence15[0:k]+'C'+new_sequence15[k+1:])
                new_sequences_4bp.append(new_sequence16[0:k]+'A'+new_sequence16[k+1:])
                new_sequences_4bp.append(new_sequence16[0:k]+'T'+new_sequence16[k+1:])
                new_sequences_4bp.append(new_sequence16[0:k]+'G'+new_sequence16[k+1:])
                new_sequences_4bp.append(new_sequence16[0:k]+'C'+new_sequence16[k+1:])
    return new_sequences_4bp


######### Where the magic happens ##########
global lbl_output
def determine_gRNAs(PAM_sequences,kill_strains,nonkill_strains,nucs_specificity,target_length2,PAM_orientation):
    global lbl_output, sleep_time, PAM_seqs, kill_strains_list, nonkill_strains_list, num_gRNAs, btn_download, gRNAs_1mismatch, next_all_gRNAs_other, all_gRNAs, all_gRNAs_nontargets, target_length, PAM_orientation_output, nucleotides_specificity
    
    kill_strains_list = kill_strains
    nonkill_strains_list = nonkill_strains
    nucleotides_specificity = nucs_specificity    
    ##### Enter list of PAM sequences #####
    user_email=str(entry_email.get())
    if user_email=='' or user_email=='Enter email address. This is only used for NCBI searches.':
        lbl_output.config(font=("Arial", 12), background="red2", text="Enter email", width=35)
        return
    Entrez.email = user_email
    PAM_sequences=str(PAM_sequences)
    if PAM_sequences=='Enter PAMs (ex: AGG, TGG, CGG, GGG)':
        lbl_output.config(font=("Arial", 12), background="red2", text="No PAM sequences entered", width=35)
        return
    PAM_sequences = PAM_sequences.replace(" ","")
    PAM_seqs=[]
    prev_pam=0
    next_pam=-1
    counter=0
    for letter in PAM_sequences:
        if letter == "," or letter == ";":
            next_pam = counter
            PAM_seqs.append(PAM_sequences[prev_pam:next_pam])
            prev_pam=next_pam+1
        counter+=1
        if counter == len(PAM_sequences):
            PAM_seqs.append(PAM_sequences[prev_pam:counter])
    for PAM in PAM_seqs:
        for letter in PAM:
            if letter not in ['A','G','C','T']:
                lbl_output.config(font=("Arial", 12), background="red2", text="Incorrect PAM format", width=35)
                return
    if nucs_specificity=='Nucleotides of specificity':
        nucs_specificity=3
    else:
        nucs_specificity=int(nucs_specificity)
    
    if target_length2=='Target site length (bp)':
        target_length=20
    else:
        target_length=int(target_length2)
        
    if PAM_orientation=='5\'-PAM-target-3\'':
        PAM_orientation=1
        PAM_orientation_output='5\'-PAM-target-3\''
    else:
        PAM_orientation=0
        PAM_orientation_output='5\'-target-PAM-3\''
            
    t1, t2, t3, t4, nuc = 0, 0, 0, 0, 0
    if len(nonkill_strains_list)>0:
        if nucs_specificity==1:
            t1=1
            nuc=1
        if nucs_specificity==2:
            t2=40
            nuc=2
        if nucs_specificity==3:
            t3=150
            nuc=3
        if nucs_specificity==4:
            t4=240
            nuc=4
    if Cpf1Check.get() == 1:
        t_thermo=80
    elif Cas9Check.get() == 1:
        t_thermo=80
    else:
        t_thermo=0
    t_full=round((t1+t2+t3+t4+t_thermo+4*len(kill_strains_list)+2.7*len(nonkill_strains_list)*nuc+31)/60)
    time_text="Identifying gRNAs: "+ str(t_full) +" min remaining"
    lbl_output.config(text=time_text,background=background_color,font=("Arial", 12))
    window.update()
    
    
    ##### Obtain genomic sequence for first selected target strain #####
    my_dna=''
    if kill_strains_list[0]=='User-defined target sequence':
        userdefined_target_file=askopenfilename(filetypes=[('fasta file','*.fa'), ('fasta file','.fasta'),('fasta file','.fna'),('fasta file','.ffn'),('fasta file','.faa'),('fasta file','.frn')], title='Select target strain Fasta file')
        userdefined_target_sequence = SeqIO.parse(userdefined_target_file, "fasta")
        for fasta in userdefined_target_sequence:
            my_dna=str(fasta.seq)
            my_dna=my_dna.upper()
    else:
        if type(all_strains_numbers_flat2[all_strains_numbers_flat2.index(kill_strains[0])+1])==list: ## This is true if the strain has multiple genomes
            for refID in all_strains_numbers_flat2[all_strains_numbers_flat2.index(kill_strains[0])+1]:
                for i in range(10):
                    try:
                        handle = Entrez.efetch(db="nucleotide", id=refID, rettype="fasta", retmode="text")
                        break
                    except:
                        time.sleep(sleep_time)
                        if i==9:
                            lbl_output.config(font=("Arial", 12), background="red2", text="Failed to connect to NCBI.", width=35)
                            return
                my_dna_new=handle.read()
                my_dna_new=my_dna_new.replace('\n','')
                start_genome=my_dna_new.find('genome')+6
                if start_genome==5:
                    start_genome=my_dna_new.find('sequence')+8
                my_dna_new=my_dna_new[start_genome:]
                my_dna_new.upper()
                my_dna=my_dna+my_dna_new
        else: # If the strain just has one genome
            for i in range(10):
                try:
                    handle = Entrez.efetch(db="nucleotide", id=all_strains_numbers_flat2[all_strains_numbers_flat2.index(kill_strains[0])+1], rettype="fasta", retmode="text")
                    break
                except:
                    time.sleep(sleep_time)
                    if i==9:
                        lbl_output.config(font=("Arial", 12), background="red2", text="Failed to connect to NCBI.", width=35)
                        return
            my_dna_new=handle.read()
            my_dna_new=my_dna_new.replace('\n','')
            start_genome=my_dna_new.find('genome')+6
            if start_genome==5:
                start_genome=my_dna_new.find('sequence')+8
            my_dna_new=my_dna_new[start_genome:]
            my_dna_new.upper()
            my_dna=my_dna+my_dna_new
            del my_dna_new
    
    my_dna_rev=Seq(my_dna,DNAAlphabet()).reverse_complement()

    ##### find locations for all PAM sites
    allnumGG1=[]
    allnumGG_rev1=[]
    for pam in PAM_seqs:
        allnumGG1.append([m.start() for m in re.finditer('(?='+str(pam)+')',str(my_dna).upper())])
        allnumGG_rev1.append([m.start() for m in re.finditer('(?='+str(pam)+')',str(my_dna_rev).upper())])
    allnumGG=[]
    allnumGG_rev=[]
    for sublist in allnumGG1:
        for item in sublist:
            allnumGG.append(item)
    for sublist in allnumGG_rev1:
        for item in sublist:
            allnumGG_rev.append(item)

    ##### Convert PAM sites to full gRNAs #####
    all_gRNAs=[]
    for n in allnumGG:
        if n>(target_length+2):
            all_gRNAs.append(my_dna[(n-target_length+PAM_orientation*(target_length+len(PAM_seqs[0]))):(n+PAM_orientation*(target_length+len(PAM_seqs[0])))])
    for n in allnumGG_rev:
        if n>(target_length+2):
            all_gRNAs.append(my_dna_rev[(n-target_length+PAM_orientation*(target_length+len(PAM_seqs[0]))):(n+PAM_orientation*(target_length+len(PAM_seqs[0])))])

    del allnumGG1, allnumGG_rev1, allnumGG, allnumGG_rev, my_dna
    
    
    ##### Only keep gRNAs with perfect matches in all strains #####
    perfect_gRNAs = []
    s=1
    while s < len(kill_strains):
        # Obtain sequence for next selected kill strain
        next_other_dna=''
        if kill_strains_list[s]=='User-defined target sequence':
            userdefined_target_file=askopenfilename(filetypes=[('fasta file','*.fa'), ('fasta file','.fasta'),('fasta file','.fna'),('fasta file','.ffn'),('fasta file','.faa'),('fasta file','.frn')], title='Select target strain Fasta file')
            userdefined_target_sequence = SeqIO.parse(userdefined_target_file, "fasta")
            for fasta in userdefined_target_sequence:
                next_other_dna=str(fasta.seq)
                next_other_dna=next_other_dna.upper()
        else:
            if type(all_strains_numbers_flat2[all_strains_numbers_flat2.index(kill_strains[s])+1])==list:
                for refID in all_strains_numbers_flat2[all_strains_numbers_flat2.index(kill_strains[s])+1]:
                    for i in range(10):
                        try:
                            handle = Entrez.efetch(db="nucleotide", id=refID, rettype="fasta", retmode="text")
                            break
                        except:
                            time.sleep(sleep_time)
                            if i==9:
                                lbl_output.config(font=("Arial", 12), background="red2", text="Failed to connect to NCBI.", width=35)
                                return
                    next_other_dna_new=handle.read()
                    next_other_dna_new=next_other_dna_new.replace('\n','')
                    start_genome=next_other_dna_new.find('genome')+6
                    if start_genome==5:
                        start_genome=next_other_dna_new.find('sequence')+8
                    next_other_dna_new=next_other_dna_new[start_genome:]
                    next_other_dna_new.upper()
                    next_other_dna=next_other_dna+next_other_dna_new
            else: # If the strain just has one genome
                for i in range(10):
                    try:
                        handle = Entrez.efetch(db="nucleotide", id=all_strains_numbers_flat2[all_strains_numbers_flat2.index(kill_strains[s])+1], rettype="fasta", retmode="text")
                        break
                    except:
                        time.sleep(sleep_time)
                        if i==9:
                            lbl_output.config(font=("Arial", 12), background="red2", text="Failed to connect to NCBI.", width=35)
                            return
                
                next_other_dna_new=handle.read()
                next_other_dna_new=next_other_dna_new.replace('\n','')
                start_genome=next_other_dna_new.find('genome')+6
                if start_genome==5:
                    start_genome=next_other_dna_new.find('sequence')+8
                next_other_dna_new=next_other_dna_new[start_genome:]
                next_other_dna_new.upper()
                next_other_dna=next_other_dna+next_other_dna_new
                del next_other_dna_new
        
        
        next_other_dna_rev=Seq(next_other_dna,DNAAlphabet()).reverse_complement()


        # Determine locations for all PAM sites in strain
        allnumGG1=[]
        allnumGG_rev1=[]
        for pam in PAM_seqs:
            allnumGG1.append([m.start() for m in re.finditer('(?='+str(pam)+')',str(next_other_dna).upper())])
            allnumGG_rev1.append([m.start() for m in re.finditer('(?='+str(pam)+')',str(next_other_dna_rev).upper())])
        allnumGG_next_other=[]
        allnumGG_next_rev_other=[]
        for sublist in allnumGG1:
            for item in sublist:
                allnumGG_next_other.append(item)
        for sublist in allnumGG_rev1:
            for item in sublist:
                allnumGG_next_rev_other.append(item)
        # Determine all gRNA target site sequences in selected strain based on PAM locations
        next_all_gRNAs_other=[]
        for n in allnumGG_next_other:
            if n>(target_length+2):
                next_all_gRNAs_other.append(str(next_other_dna[(n-target_length+PAM_orientation*(target_length+len(PAM_seqs[0]))):(n+PAM_orientation*(target_length+len(PAM_seqs[0])))]))
        for n in allnumGG_next_rev_other:
            if n>(target_length+2):
                next_all_gRNAs_other.append(str(next_other_dna_rev[(n-target_length+PAM_orientation*(target_length+len(PAM_seqs[0]))):(n+PAM_orientation*(target_length+len(PAM_seqs[0])))]))
        # Determine which gRNA sequences from initial list are also present in this strain with 100% match

        perfect_gRNAs = list(set(next_all_gRNAs_other).intersection(set(all_gRNAs)))

        if len(perfect_gRNAs)==0:
            lbl_output.config(font=("Arial", 12), background="red2", text="This selection has 0 possible gRNAs.", width=35)
            return


        all_gRNAs=perfect_gRNAs
        perfect_gRNAs=[]

        if s%15==0:
            t_full=round((t1+t2+t3+t4+t_thermo+4*(len(kill_strains_list)-s)+2.7*len(nonkill_strains_list)*nuc+31)/60)
            time_text="Identifying gRNAs: "+ str(t_full) +" min remaining"
            lbl_output.config(text=time_text,background=background_color,font=("Arial", 12))
            window.update()

        del next_other_dna, next_other_dna_rev, allnumGG_next_other, allnumGG_next_rev_other, perfect_gRNAs, next_all_gRNAs_other
        s = s+1
    
    
    ###########################################################################################################################
    ############# Determine gRNAs that avoid targeting all selected nontarget strains using 2bp of specificity ################
    ###########################################################################################################################
    
    s=0
    if len(nonkill_strains)>0:
        user_defined_seqs=[]
        for pos25 in range(int(math.ceil(len(nonkill_strains) / 25.0))):
            nontarget_dna_total=''
            if pos25 == int(math.ceil(len(nonkill_strains) / 25.0))-1:
                num_strains = len(nonkill_strains)-25*(int(math.ceil(len(nonkill_strains) / 25.0))-1)
            else:
                num_strains = 25
                
            for iter_number in range(num_strains):
                # Obtain sequences for all nontarget strains
                if nonkill_strains_list[s]=='User-defined non-target sequence':
                    userdefined_nontarget_file=askopenfilename(filetypes=[('fasta file','*.fa'), ('fasta file','.fasta'),('fasta file','.fna'),('fasta file','.ffn'),('fasta file','.faa'),('fasta file','.frn')], title='Select non-target strain Fasta file')
                    userdefined_nontarget_sequence = SeqIO.parse(userdefined_nontarget_file, "fasta")
                    for fasta in userdefined_nontarget_sequence:
                        next_other_dna_new=str(fasta.seq)
                        next_other_dna_new=next_other_dna_new.upper()
                        user_defined_seqs.append(next_other_dna_new)
                        nontarget_dna_total=nontarget_dna_total+next_other_dna_new
                else:
                    if type(all_strains_numbers_flat2[all_strains_numbers_flat2.index(nonkill_strains[s])+1])==list:
                        for refID in all_strains_numbers_flat2[all_strains_numbers_flat2.index(nonkill_strains[s])+1]:
                            for i in range(10):
                                try:
                                    handle = Entrez.efetch(db="nucleotide", id=refID, rettype="fasta", retmode="text")
                                    break
                                except:
                                    time.sleep(sleep_time)
                                    if i==9:
                                        lbl_output.config(font=("Arial", 12), background="red2", text="Failed to connect to NCBI.", width=35)
                                        return
                            next_other_dna_new=handle.read()
                            next_other_dna_new=next_other_dna_new.replace('\n','')
                            start_genome=next_other_dna_new.find('genome')+6
                            if start_genome==5:
                                start_genome=next_other_dna_new.find('sequence')+8
                            next_other_dna_new=next_other_dna_new[start_genome:]
                            next_other_dna_new.upper()
                            nontarget_dna_total=nontarget_dna_total+next_other_dna_new
                    else: # If the strain just has one genome
                        for i in range(10):
                            try:
                                handle = Entrez.efetch(db="nucleotide", id=all_strains_numbers_flat2[all_strains_numbers_flat2.index(nonkill_strains[s])+1], rettype="fasta", retmode="text")
                                break
                            except:
                                time.sleep(sleep_time)
                                if i==9:
                                    lbl_output.config(font=("Arial", 12), background="red2", text="Failed to connect to NCBI.", width=35)
                                    return
                        next_other_dna_new=handle.read()
                        next_other_dna_new=next_other_dna_new.replace('\n','')
                        start_genome=next_other_dna_new.find('genome')+6
                        if start_genome==5:
                            start_genome=next_other_dna_new.find('sequence')+8
                        next_other_dna_new=next_other_dna_new[start_genome:]
                        next_other_dna_new.upper()
                        nontarget_dna_total=nontarget_dna_total+next_other_dna_new
                    del next_other_dna_new, handle, start_genome
                s += 1
            
        
            nontarget_dna_total_rev=Seq(str(nontarget_dna_total),DNAAlphabet()).reverse_complement()

            #### Determine all nontarget gRNAs ####
    
            # Determine locations for all PAM sites in strain
            allnumGG1=[]
            allnumGG_rev1=[]
            for pam in PAM_seqs:
                allnumGG1.append([m.start() for m in re.finditer('(?='+str(pam)+')',str(nontarget_dna_total).upper())])
                allnumGG_rev1.append([m.start() for m in re.finditer('(?='+str(pam)+')',str(nontarget_dna_total_rev).upper())])
            allnumGG_next_other=[]
            allnumGG_next_rev_other=[]
            for sublist in allnumGG1:
                for item in sublist:
                    allnumGG_next_other.append(item)
            for sublist in allnumGG_rev1:
                for item in sublist:
                    allnumGG_next_rev_other.append(item)
            # Determine all gRNA target site sequences in selected strain based on PAM locations
            # Only take the 10bp immediately adjascent to the PAM site (most important)
            all_gRNAs_nontargets=[]
            for n in allnumGG_next_other:
                if n>(target_length+2):
                    all_gRNAs_nontargets.append(str(nontarget_dna_total[(n-target_length+PAM_orientation*(target_length+len(PAM_seqs[0]))):(n+PAM_orientation*(target_length+len(PAM_seqs[0])))]))
            for n in allnumGG_next_rev_other:
                if n>(target_length+2):
                    all_gRNAs_nontargets.append(str(nontarget_dna_total_rev[(n-target_length+PAM_orientation*(target_length+len(PAM_seqs[0]))):(n+PAM_orientation*(target_length+len(PAM_seqs[0])))]))
                    
            del allnumGG1, allnumGG_rev1, allnumGG_next_other, allnumGG_next_rev_other
    
            all_gRNAs_nontargets = list(dict.fromkeys(all_gRNAs_nontargets))
            all_gRNAs = list(dict.fromkeys(all_gRNAs))
            
            # Remove exact matches between target and nontarget strains
            
            remove_gRNAs = list(set(all_gRNAs).intersection(set(all_gRNAs_nontargets)))
            all_gRNAs = list(set(all_gRNAs)-set(remove_gRNAs))
            
            if len(all_gRNAs)==0:
                lbl_output.config(font=("Arial", 12), background="red2", text="This selection has 0 possible gRNAs.", width=35)
                return
            
            del remove_gRNAs
            
            if s<26 or nucs_specificity==1:
                t_full=round((t2+t3+t4+t_thermo+2.7*len(nonkill_strains_list)*(nuc-s/len(nonkill_strains_list))+31)/60)
                time_text="Identifying gRNAs: "+ str(t_full) +" min remaining"
                lbl_output.config(text=time_text,background=background_color,font=("Arial", 12))
                window.update()
            
            # Check whether gRNAs for target strains are also present in any nontarget strain. If present, remove gRNA
            # Create 1bp mismatches in the target strain gRNAs
            if nucs_specificity>1:
                num_gRNAs = len(all_gRNAs)
                if len(all_gRNAs)>max_gRNAs_2bp:
                    gRNAs_subset=random.sample(all_gRNAs,max_gRNAs_2bp)
                    all_gRNAs=gRNAs_subset
                
                if pos25==0:
                    gRNAs_1mismatch=list(map(replace_nucs,all_gRNAs))


                all_gRNAs_check=list(map(set(all_gRNAs_nontargets).isdisjoint,gRNAs_1mismatch))

                all_gRNAs=list(compress(all_gRNAs,all_gRNAs_check))
                gRNAs_1mismatch=list(compress(gRNAs_1mismatch,all_gRNAs_check))
                
                del all_gRNAs_check
            
                t_full=round((t3+t4+t_thermo+2.7*len(nonkill_strains_list)*(nuc-1-s/len(nonkill_strains_list))+31)/60)
                time_text="Identifying gRNAs: "+ str(t_full) +" min remaining"
                lbl_output.config(text=time_text,background=background_color,font=("Arial", 12))
                window.update()
            
            if len(all_gRNAs)==0:
                lbl_output.config(font=("Arial", 12), background="red2", text="This selection has 0 possible gRNAs.", width=35)
                return

        del gRNAs_1mismatch
        ###########################################################################################################################
        ############# Determine gRNAs that avoid targeting all selected nontarget strains using 3bp of specificity ################
        ###########################################################################################################################
        if nucs_specificity>2:
            #Cap the number of gRNAs to test here
            num_gRNAs = len(all_gRNAs)
            if len(all_gRNAs)>max_gRNAs_3bp:
                gRNAs_subset=random.sample(all_gRNAs,max_gRNAs_3bp)
                all_gRNAs=gRNAs_subset
            all_gRNAs_final = []
            for i in all_gRNAs:
                all_gRNAs_final.append(i.upper())
            all_gRNAs = all_gRNAs_final
            
            # Create 2bp mismatches in the target strain gRNAs
            gRNAs_2mismatch=list(map(replace_nucs2bp,all_gRNAs))
            
            t_full=round((t4+t_thermo+2.7*len(nonkill_strains_list)*(nuc-2)+31)/60)
            time_text="Identifying gRNAs: "+ str(t_full) +" min remaining"
            lbl_output.config(text=time_text,background=background_color,font=("Arial", 12))
            window.update()
            
            num_user=0
            s=0
            for pos25 in range(int(math.ceil(len(nonkill_strains) / 25.0))):
                nontarget_dna_total=''
                if pos25 == int(math.ceil(len(nonkill_strains) / 25.0))-1:
                    num_strains = len(nonkill_strains)-25*(int(math.ceil(len(nonkill_strains) / 25.0))-1)
                else:
                    num_strains = 25
                    
                for iter_number in range(num_strains):
                    # Obtain sequences for all nontarget strains
                    if nonkill_strains_list[s]=='User-defined non-target sequence':
                        next_other_dna_new=user_defined_seqs[num_user]
                        next_other_dna_new.upper()
                        nontarget_dna_total=nontarget_dna_total+next_other_dna_new
                        num_user+=1
                    else:
                        if type(all_strains_numbers_flat2[all_strains_numbers_flat2.index(nonkill_strains[s])+1])==list:
                            for refID in all_strains_numbers_flat2[all_strains_numbers_flat2.index(nonkill_strains[s])+1]:
                                for i in range(10):
                                    try:
                                        handle = Entrez.efetch(db="nucleotide", id=refID, rettype="fasta", retmode="text")
                                        break
                                    except:
                                        time.sleep(sleep_time)
                                        if i==9:
                                            lbl_output.config(font=("Arial", 12), background="red2", text="Failed to connect to NCBI.", width=35)
                                            return
                                next_other_dna_new=handle.read()
                                next_other_dna_new=next_other_dna_new.replace('\n','')
                                start_genome=next_other_dna_new.find('genome')+6
                                if start_genome==5:
                                    start_genome=next_other_dna_new.find('sequence')+8
                                next_other_dna_new=next_other_dna_new[start_genome:]
                                next_other_dna_new.upper()
                                nontarget_dna_total=nontarget_dna_total+next_other_dna_new
                        else: # If the strain just has one genome
                            for i in range(10):
                                try:
                                    handle = Entrez.efetch(db="nucleotide", id=all_strains_numbers_flat2[all_strains_numbers_flat2.index(nonkill_strains[s])+1], rettype="fasta", retmode="text")
                                    break
                                except:
                                    time.sleep(sleep_time)
                                    if i==9:
                                        lbl_output.config(font=("Arial", 12), background="red2", text="Failed to connect to NCBI.", width=35)
                                        return
                            next_other_dna_new=handle.read()
                            next_other_dna_new=next_other_dna_new.replace('\n','')
                            start_genome=next_other_dna_new.find('genome')+6
                            if start_genome==5:
                                start_genome=next_other_dna_new.find('sequence')+8
                            next_other_dna_new=next_other_dna_new[start_genome:]
                            next_other_dna_new.upper()
                            nontarget_dna_total=nontarget_dna_total+next_other_dna_new
                        del next_other_dna_new, handle, start_genome
                    s += 1
                
            
                nontarget_dna_total_rev=Seq(str(nontarget_dna_total),DNAAlphabet()).reverse_complement()
                
                #### Determine all nontarget gRNAs ####
        
                # Determine locations for all PAM sites in strain
                allnumGG1=[]
                allnumGG_rev1=[]
                for pam in PAM_seqs:
                    allnumGG1.append([m.start() for m in re.finditer('(?='+str(pam)+')',str(nontarget_dna_total).upper())])
                    allnumGG_rev1.append([m.start() for m in re.finditer('(?='+str(pam)+')',str(nontarget_dna_total_rev).upper())])
                allnumGG_next_other=[]
                allnumGG_next_rev_other=[]
                for sublist in allnumGG1:
                    for item in sublist:
                        allnumGG_next_other.append(item)
                for sublist in allnumGG_rev1:
                    for item in sublist:
                        allnumGG_next_rev_other.append(item)
                # Determine all gRNA target site sequences in selected strain based on PAM locations
                # Only take the 10bp immediately adjascent to the PAM site (most important)
                all_gRNAs_nontargets=[]
                for n in allnumGG_next_other:
                    if n>(target_length+2):
                        all_gRNAs_nontargets.append(str(nontarget_dna_total[(n-target_length+PAM_orientation*(target_length+len(PAM_seqs[0]))):(n+PAM_orientation*(target_length+len(PAM_seqs[0])))]))
                for n in allnumGG_next_rev_other:
                    if n>(target_length+2):
                        all_gRNAs_nontargets.append(str(nontarget_dna_total_rev[(n-target_length+PAM_orientation*(target_length+len(PAM_seqs[0]))):(n+PAM_orientation*(target_length+len(PAM_seqs[0])))]))
                        
                del allnumGG1, allnumGG_rev1, allnumGG_next_other, allnumGG_next_rev_other
        
                all_gRNAs_nontargets = list(dict.fromkeys(all_gRNAs_nontargets))
                
                # Check whether gRNAs for target strains are also present in any nontarget strain. If present, remove gRNA
                all_gRNAs_check=list(map(set(all_gRNAs_nontargets).isdisjoint,gRNAs_2mismatch))
                all_gRNAs=list(compress(all_gRNAs,all_gRNAs_check))
                gRNAs_2mismatch=list(compress(gRNAs_2mismatch,all_gRNAs_check))
                
                del all_gRNAs_check
            
                t_full=round((t4+t_thermo+2.7*len(nonkill_strains_list)*(nuc-2-s/len(nonkill_strains_list))+31)/60)
                time_text="Identifying gRNAs: "+ str(t_full) +" min remaining"
                lbl_output.config(text=time_text,background=background_color,font=("Arial", 12))
                window.update()
                
                if len(all_gRNAs)==0:
                    lbl_output.config(font=("Arial", 12), background="red2", text="This selection has 0 possible gRNAs.", width=35)
                    return
                
            del gRNAs_2mismatch
        
    
    
    ################## CAP NUMBER OF GRNAS HERE #######################

    num_gRNAs = len(all_gRNAs)
    if len(all_gRNAs)>max_gRNAs:
        gRNAs_subset=random.sample(all_gRNAs,max_gRNAs)
        all_gRNAs=gRNAs_subset
    all_gRNAs_final = []
    for i in all_gRNAs:
        all_gRNAs_final.append(i.upper())
    all_gRNAs = all_gRNAs_final
        
    ## Remove gRNAs with non A/T/G/C characters
    remove_gs=[]
    numgs=len(all_gRNAs)
    for g in range(numgs):
        gRNA=all_gRNAs[g]
        if len(gRNA) == target_length:
            numATGCU=0
            valid_characters=['T','A','G','C']
            for i in range(0,target_length):
                if gRNA[i] in valid_characters:
                    numATGCU=numATGCU+1
            if numATGCU<target_length:
                remove_gs.append(g)
            elif numATGCU>target_length:
                remove_gs.append(g)
        else:
            remove_gs.append(g)
    
    for g in remove_gs[::-1]:
        all_gRNAs.remove(all_gRNAs[g])

    ######################################################################################################
    ################### Predict the efficiency of the gRNAs based on the properties ######################
    ######################################################################################################
    
    if Cpf1Check.get() == 1:
        all_gRNA_properties=list(map(property_identifier_Cpf1,all_gRNAs))
        model_name='finalized_model_cpf1.sav'
        model_path = os.path.join(application_path, model_name)
        prediction_model = pickle.load(open(model_path,'rb'))
        gRNA_efficiencies=prediction_model.predict(all_gRNA_properties)
        del all_gRNA_properties
        #### Sort the gRNAs by their predicted efficiency scores
        gRNAs_and_efficiencies=[]
        for j in range(len(gRNA_efficiencies)):
            gRNAs_and_efficiencies.append([all_gRNAs[j],gRNA_efficiencies[j]])
        gRNAs_and_efficiencies.sort(key=lambda x: x[1])
        all_gRNAs=[item[0] for item in gRNAs_and_efficiencies]
    elif Cas9Check.get() == 1:
        all_gRNA_properties=list(map(property_identifier_Cas9,all_gRNAs))
        model_name='finalized_model_cas9.sav'
        model_path = os.path.join(application_path, model_name)
        prediction_model = pickle.load(open(model_path,'rb'))
        gRNA_efficiencies=prediction_model.predict(all_gRNA_properties)
        del all_gRNA_properties
        #### Sort the gRNAs by their predicted efficiency scores
        gRNAs_and_efficiencies=[]
        for j in range(len(gRNA_efficiencies)):
            gRNAs_and_efficiencies.append([all_gRNAs[j],gRNA_efficiencies[j]])
        gRNAs_and_efficiencies.sort(key=lambda x: x[1])
        all_gRNAs=[item[0] for item in gRNAs_and_efficiencies]

    t_full=round((t4+2.7*len(nonkill_strains_list)*(nuc-3)+31)/60)
    time_text="Identifying gRNAs: "+ str(t_full) +" min remaining"
    lbl_output.config(text=time_text,background=background_color,font=("Arial", 12))
    window.update()
    
    ###########################################################################################################################
    ############# Determine gRNAs that avoid targeting all selected nontarget strains using 3bp of specificity ################
    ###########################################################################################################################
    if len(nonkill_strains)>0 and nucs_specificity>3:
        #Cap the number of gRNAs to test here
        num_gRNAs = len(all_gRNAs)
        if len(all_gRNAs)>max_gRNAs_4bp:
            all_gRNAs=all_gRNAs[0:max_gRNAs_4bp]
        
        # Create 3bp mismatches in the target strain gRNAs
        gRNAs_3mismatch=list(map(replace_nucs3bp,all_gRNAs))
        
        t_full=round((2*len(nonkill_strains_list)*(nuc-3)+31)/60)
        time_text="Identifying gRNAs: "+ str(t_full) +" min remaining"
        lbl_output.config(text=time_text,background=background_color,font=("Arial", 12))
        window.update()
        
        num_user=0
        s=0
        for pos25 in range(int(math.ceil(len(nonkill_strains) / 25.0))):
            nontarget_dna_total=''
            if pos25 == int(math.ceil(len(nonkill_strains) / 25.0))-1:
                num_strains = len(nonkill_strains)-25*(int(math.ceil(len(nonkill_strains) / 25.0))-1)
            else:
                num_strains = 25
                
            for iter_number in range(num_strains):
                # Obtain sequences for all nontarget strains
                if nonkill_strains_list[s]=='User-defined non-target sequence':
                    next_other_dna_new=user_defined_seqs[num_user]
                    next_other_dna_new.upper()
                    nontarget_dna_total=nontarget_dna_total+next_other_dna_new
                    num_user+=1
                else:
                    if type(all_strains_numbers_flat2[all_strains_numbers_flat2.index(nonkill_strains[s])+1])==list:
                        for refID in all_strains_numbers_flat2[all_strains_numbers_flat2.index(nonkill_strains[s])+1]:
                            for i in range(10):
                                try:
                                    handle = Entrez.efetch(db="nucleotide", id=refID, rettype="fasta", retmode="text")
                                    break
                                except:
                                    time.sleep(sleep_time)
                                    if i==9:
                                        lbl_output.config(font=("Arial", 12), background="red2", text="Failed to connect to NCBI.", width=35)
                                        return
                            next_other_dna_new=handle.read()
                            next_other_dna_new=next_other_dna_new.replace('\n','')
                            start_genome=next_other_dna_new.find('genome')+6
                            if start_genome==5:
                                start_genome=next_other_dna_new.find('sequence')+8
                            next_other_dna_new=next_other_dna_new[start_genome:]
                            next_other_dna_new.upper()
                            nontarget_dna_total=nontarget_dna_total+next_other_dna_new
                    else: # If the strain just has one genome
                        for i in range(10):
                            try:
                                handle = Entrez.efetch(db="nucleotide", id=all_strains_numbers_flat2[all_strains_numbers_flat2.index(nonkill_strains[s])+1], rettype="fasta", retmode="text")
                                break
                            except:
                                time.sleep(sleep_time)
                                if i==9:
                                    lbl_output.config(font=("Arial", 12), background="red2", text="Failed to connect to NCBI.", width=35)
                                    return
                        next_other_dna_new=handle.read()
                        next_other_dna_new=next_other_dna_new.replace('\n','')
                        start_genome=next_other_dna_new.find('genome')+6
                        if start_genome==5:
                            start_genome=next_other_dna_new.find('sequence')+8
                        next_other_dna_new=next_other_dna_new[start_genome:]
                        next_other_dna_new.upper()
                        nontarget_dna_total=nontarget_dna_total+next_other_dna_new
                    del next_other_dna_new, handle, start_genome
                s += 1
            
        
            nontarget_dna_total_rev=Seq(str(nontarget_dna_total),DNAAlphabet()).reverse_complement()

            #### Determine all nontarget gRNAs ####
    
            # Determine locations for all PAM sites in strain
            allnumGG1=[]
            allnumGG_rev1=[]
            for pam in PAM_seqs:
                allnumGG1.append([m.start() for m in re.finditer('(?='+str(pam)+')',str(nontarget_dna_total).upper())])
                allnumGG_rev1.append([m.start() for m in re.finditer('(?='+str(pam)+')',str(nontarget_dna_total_rev).upper())])
            allnumGG_next_other=[]
            allnumGG_next_rev_other=[]
            for sublist in allnumGG1:
                for item in sublist:
                    allnumGG_next_other.append(item)
            for sublist in allnumGG_rev1:
                for item in sublist:
                    allnumGG_next_rev_other.append(item)
            # Determine all gRNA target site sequences in selected strain based on PAM locations
            # Only take the 10bp immediately adjascent to the PAM site (most important)
            all_gRNAs_nontargets=[]
            for n in allnumGG_next_other:
                if n>(target_length+2):
                    all_gRNAs_nontargets.append(str(nontarget_dna_total[(n-target_length+PAM_orientation*(target_length+len(PAM_seqs[0]))):(n+PAM_orientation*(target_length+len(PAM_seqs[0])))]))
            for n in allnumGG_next_rev_other:
                if n>(target_length+2):
                    all_gRNAs_nontargets.append(str(nontarget_dna_total_rev[(n-target_length+PAM_orientation*(target_length+len(PAM_seqs[0]))):(n+PAM_orientation*(target_length+len(PAM_seqs[0])))]))
                    
            del allnumGG1, allnumGG_rev1, allnumGG_next_other, allnumGG_next_rev_other
    
            all_gRNAs_nontargets = list(dict.fromkeys(all_gRNAs_nontargets))
            
            # Check whether gRNAs for target strains are also present in any nontarget strain. If present, remove gRNA
            all_gRNAs_check=list(map(set(all_gRNAs_nontargets).isdisjoint,gRNAs_3mismatch))
            all_gRNAs=list(compress(all_gRNAs,all_gRNAs_check))
            gRNAs_3mismatch=list(compress(gRNAs_3mismatch,all_gRNAs_check))
            
            del all_gRNAs_check
            
            t_full=round((2*len(nonkill_strains_list)*(nuc-3-1*s/len(nonkill_strains))+31)/60)
            time_text="Identifying gRNAs: "+ str(t_full) +" min remaining"
            lbl_output.config(text=time_text,background=background_color,font=("Arial", 12))
            window.update()
            
            if len(all_gRNAs)==0:
                lbl_output.config(font=("Arial", 12), background="red2", text="This selection has 0 possible gRNAs.", width=35)
                return
            
        del gRNAs_3mismatch
    
    num_gRNAs = len(all_gRNAs)

    ##### Output results of the gRNA search #####
    lbl_output.config(font=("Arial", 12), background="pale green", text="This selection has "+str(num_gRNAs)+" possible gRNAs", width=35)
    btn_download = Button(window, text = "Download results", command = download_results)
    btn_download.config(font=("Arial", 12), width=15, background="royal blue")
    btn_download.grid(row=16,column=20,columnspan=10)


##### Efficiency for Cas9 gRNAs with 5'-Target-PAM-3' orientation #####
def property_identifier_Cas9(gRNA):
    gRNA_properties=[0]*396

    "Check num of Ts in -4 to -1 from PAM"
    numTs_14=0
    for j in [target_length-4,target_length-3,target_length-2,target_length-1]:
        if gRNA[j]=='T':
            numTs_14=numTs_14+1
    gRNA_properties[0]=numTs_14

    "Determine GC content"
    numGCs=0
    for j in range(0,target_length):
        if gRNA[j]=='G' or gRNA[j]=='C':
            numGCs=numGCs+1
    gRNA_properties[1]=numGCs/20

    "Determine A content"
    numAs=0
    for j in range(0,target_length):
        if gRNA[j]=='A':
            numAs=numAs+1
    gRNA_properties[2]=numAs/20

    "Determine T content"
    numTs=0
    for j in range(0,target_length):
        if gRNA[j]=='T':
            numTs=numTs+1
    gRNA_properties[3]=numTs/20

    "Determine G content"
    numGs=0
    for j in range(0,target_length):
        if gRNA[j]=='G':
            numGs=numGs+1
    gRNA_properties[4]=numGs/20

    "Determine C content"
    numCs=0
    for j in range(0,target_length):
        if gRNA[j]=='C':
            numCs=numCs+1
    gRNA_properties[5]=numCs/20

    "Determine free energy of full RNA sequence"
    gRNA_RNA=str(gRNA).replace('T','U')

    if dg(gRNA_RNA, temp = 37.0) > -100:
        if dg(gRNA_RNA, temp = 37.0) < 4:
            gRNA_properties[6]=dg(gRNA_RNA, temp = 37.0)
        else:
            gRNA_properties[6]=4
    else:
        gRNA_properties[6]=4


    "Determine free energy of the 12bp of RNA closest to the PAM"
    gRNA_RNA_short=gRNA_RNA[target_length-12:target_length]
    if dg(gRNA_RNA_short, temp = 37.0) > -30:
        if dg(gRNA_RNA_short, temp = 37.0) < 4:
            gRNA_properties[7]=dg(gRNA_RNA_short, temp = 37.0)
        elif str(abs(dg(gRNA_RNA_short, temp = 37.0))) == 'inf':
            gRNA_properties[7]=4
        else:
            gRNA_properties[7]=4
    else:
        gRNA_properties[7]=4

    "Calculate melting temp for full gRNA"
    gRNA_properties[8]=mt.Tm_NN(Seq(str(gRNA)).complement(), c_seq=Seq(gRNA_RNA))
    "Calculate melting temp for 5bp closest to PAM"
    gRNA_properties[9]=mt.Tm_NN(Seq(str(gRNA)).complement()[target_length-5:target_length], c_seq=Seq(gRNA_RNA)[target_length-5:target_length])
    "Calculate melting temp for middle 8bp"
    gRNA_properties[10]=mt.Tm_NN(Seq(str(gRNA)).complement()[target_length-13:target_length-5], c_seq=Seq(gRNA_RNA)[target_length-13:target_length-5])
    "Calculate melting temp for remaining bp"
    gRNA_properties[11]=mt.Tm_NN(Seq(str(gRNA)).complement()[0:target_length-13], c_seq=Seq(gRNA_RNA)[0:target_length-13])

    "Determine whether an A is in each position of 20bp near PAM"
    for j in range(target_length-20,target_length):
        if gRNA[j]=='A':
            gRNA_properties[12+j-(target_length-20)]=1
        else:
            gRNA_properties[12+j-(target_length-20)]=0

    "Determine whether a T is in each position"
    for j in range(target_length-20,target_length):
        if gRNA[j]=='T':
            gRNA_properties[32+j-(target_length-20)]=1
        else:
            gRNA_properties[32+j-(target_length-20)]=0

    "Determine whether a G is in each position"
    for j in range(target_length-20,target_length):
        if gRNA[j]=='G':
            gRNA_properties[52+j-(target_length-20)]=1
        else:
            gRNA_properties[52+j-(target_length-20)]=0

    "Determine whether a C is in each position"
    for j in range(target_length-20,target_length):
        if gRNA[j]=='C':
            gRNA_properties[72+j-(target_length-20)]=1
        else:
            gRNA_properties[72+j-(target_length-20)]=0

    "Determine whether AA is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='AA':
            gRNA_properties[92+j-(target_length-20)]=1
        else:
            gRNA_properties[92+j-(target_length-20)]=0

    "Determine whether AT is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='AT':
            gRNA_properties[111+j-(target_length-20)]=1
        else:
            gRNA_properties[111+j-(target_length-20)]=0

    "Determine whether AG is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='AG':
            gRNA_properties[130+j-(target_length-20)]=1
        else:
            gRNA_properties[130+j-(target_length-20)]=0

    "Determine whether AC is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='AC':
            gRNA_properties[149+j-(target_length-20)]=1
        else:
            gRNA_properties[149+j-(target_length-20)]=0

    "Determine whether TA is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='TA':
            gRNA_properties[168+j-(target_length-20)]=1
        else:
            gRNA_properties[168+j-(target_length-20)]=0

    "Determine whether TT is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='TT':
            gRNA_properties[187+j-(target_length-20)]=1
        else:
            gRNA_properties[187+j-(target_length-20)]=0

    "Determine whether TG is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='TG':
            gRNA_properties[206+j-(target_length-20)]=1
        else:
            gRNA_properties[206+j-(target_length-20)]=0

    "Determine whether TC is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='TC':
            gRNA_properties[225+j-(target_length-20)]=1
        else:
            gRNA_properties[225+j-(target_length-20)]=0

    "Determine whether GA is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='GA':
            gRNA_properties[244+j-(target_length-20)]=1
        else:
            gRNA_properties[244+j-(target_length-20)]=0

    "Determine whether GT is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='GT':
            gRNA_properties[263+j-(target_length-20)]=1
        else:
            gRNA_properties[263+j-(target_length-20)]=0

    "Determine whether GG is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='GG':
            gRNA_properties[282+j-(target_length-20)]=1
        else:
            gRNA_properties[282+j-(target_length-20)]=0

    "Determine whether GC is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='GC':
            gRNA_properties[301+j-(target_length-20)]=1
        else:
            gRNA_properties[301+j-(target_length-20)]=0

    "Determine whether CA is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='CA':
            gRNA_properties[320+j-(target_length-20)]=1
        else:
            gRNA_properties[320+j-(target_length-20)]=0

    "Determine whether CT is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='CT':
            gRNA_properties[339+j-(target_length-20)]=1
        else:
            gRNA_properties[339+j-(target_length-20)]=0

    "Determine whether CG is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='CG':
            gRNA_properties[358+j-(target_length-20)]=1
        else:
            gRNA_properties[358+j-(target_length-20)]=0

    "Determine whether CC is in each 2x position"
    for j in range(target_length-20,target_length-1):
        if gRNA[j:j+2]=='CC':
            gRNA_properties[376+j-(target_length-20)]=1
        else:
            gRNA_properties[376+j-(target_length-20)]=0

    return gRNA_properties

##### Efficiency for Cpf1 gRNAs with 5'-PAM-Target-3' orientation #####
def property_identifier_Cpf1(gRNA):
    gRNA_properties=[0]*396

    "Check num of Ts in -4 to -1 from PAM"
    numTs_14=0
    for j in [0,1,2,3]:
        if gRNA[j]=='T':
            numTs_14=numTs_14+1
    gRNA_properties[0]=numTs_14

    "Determine GC content"
    numGCs=0
    for j in range(0,target_length):
        if gRNA[j]=='G' or gRNA[j]=='C':
            numGCs=numGCs+1
    gRNA_properties[1]=numGCs/20

    "Determine A content"
    numAs=0
    for j in range(0,target_length):
        if gRNA[j]=='A':
            numAs=numAs+1
    gRNA_properties[2]=numAs/20

    "Determine T content"
    numTs=0
    for j in range(0,target_length):
        if gRNA[j]=='T':
            numTs=numTs+1
    gRNA_properties[3]=numTs/20

    "Determine G content"
    numGs=0
    for j in range(0,target_length):
        if gRNA[j]=='G':
            numGs=numGs+1
    gRNA_properties[4]=numGs/20

    "Determine C content"
    numCs=0
    for j in range(0,target_length):
        if gRNA[j]=='C':
            numCs=numCs+1
    gRNA_properties[5]=numCs/20

    "Determine free energy of full RNA sequence"
    gRNA_RNA=str(gRNA).replace('T','U')

    if dg(gRNA_RNA, temp = 37.0) > -100:
        if dg(gRNA_RNA, temp = 37.0) < 4:
            gRNA_properties[6]=dg(gRNA_RNA, temp = 37.0)
        else:
            gRNA_properties[6]=4
    else:
        gRNA_properties[6]=4


    "Determine free energy of the 12bp of RNA closest to the PAM"
    gRNA_RNA_short=gRNA_RNA[0:12]
    if dg(gRNA_RNA_short, temp = 37.0) > -30:
        if dg(gRNA_RNA_short, temp = 37.0) < 4:
            gRNA_properties[7]=dg(gRNA_RNA_short, temp = 37.0)
        elif str(abs(dg(gRNA_RNA_short, temp = 37.0))) == 'inf':
            gRNA_properties[7]=4
        else:
            gRNA_properties[7]=4
    else:
        gRNA_properties[7]=4

    "Calculate melting temp for full gRNA"
    gRNA_properties[8]=mt.Tm_NN(Seq(str(gRNA)).complement(), c_seq=Seq(gRNA_RNA))
    "Calculate melting temp for 5bp closest to PAM"
    gRNA_properties[9]=mt.Tm_NN(Seq(str(gRNA)).complement()[0:5], c_seq=Seq(gRNA_RNA)[0:5])
    "Calculate melting temp for middle 8bp"
    gRNA_properties[10]=mt.Tm_NN(Seq(str(gRNA)).complement()[5:13], c_seq=Seq(gRNA_RNA)[5:13])
    "Calculate melting temp for remaining bp"
    gRNA_properties[11]=mt.Tm_NN(Seq(str(gRNA)).complement()[13:target_length], c_seq=Seq(gRNA_RNA)[13:target_length])

    "Determine whether an A is in each position of 20bp near PAM"
    for j in range(19,-1,-1):
        if gRNA[j]=='A':
            gRNA_properties[12+j-(target_length-20)]=1
        else:
            gRNA_properties[12+j-(target_length-20)]=0

    "Determine whether a T is in each position"
    for j in range(19,-1,-1):
        if gRNA[j]=='T':
            gRNA_properties[32+j-(target_length-20)]=1
        else:
            gRNA_properties[32+j-(target_length-20)]=0

    "Determine whether a G is in each position"
    for j in range(19,-1,-1):
        if gRNA[j]=='G':
            gRNA_properties[52+j-(target_length-20)]=1
        else:
            gRNA_properties[52+j-(target_length-20)]=0

    "Determine whether a C is in each position"
    for j in range(19,-1,-1):
        if gRNA[j]=='C':
            gRNA_properties[72+j-(target_length-20)]=1
        else:
            gRNA_properties[72+j-(target_length-20)]=0

    "Determine whether AA is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='AA':
            gRNA_properties[92+j-(target_length-20)]=1
        else:
            gRNA_properties[92+j-(target_length-20)]=0

    "Determine whether AT is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='AT':
            gRNA_properties[111+j-(target_length-20)]=1
        else:
            gRNA_properties[111+j-(target_length-20)]=0

    "Determine whether AG is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='AG':
            gRNA_properties[130+j-(target_length-20)]=1
        else:
            gRNA_properties[130+j-(target_length-20)]=0

    "Determine whether AC is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='AC':
            gRNA_properties[149+j-(target_length-20)]=1
        else:
            gRNA_properties[149+j-(target_length-20)]=0

    "Determine whether TA is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='TA':
            gRNA_properties[168+j-(target_length-20)]=1
        else:
            gRNA_properties[168+j-(target_length-20)]=0

    "Determine whether TT is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='TT':
            gRNA_properties[187+j-(target_length-20)]=1
        else:
            gRNA_properties[187+j-(target_length-20)]=0

    "Determine whether TG is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='TG':
            gRNA_properties[206+j-(target_length-20)]=1
        else:
            gRNA_properties[206+j-(target_length-20)]=0

    "Determine whether TC is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='TC':
            gRNA_properties[225+j-(target_length-20)]=1
        else:
            gRNA_properties[225+j-(target_length-20)]=0

    "Determine whether GA is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='GA':
            gRNA_properties[244+j-(target_length-20)]=1
        else:
            gRNA_properties[244+j-(target_length-20)]=0

    "Determine whether GT is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='GT':
            gRNA_properties[263+j-(target_length-20)]=1
        else:
            gRNA_properties[263+j-(target_length-20)]=0

    "Determine whether GG is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='GG':
            gRNA_properties[282+j-(target_length-20)]=1
        else:
            gRNA_properties[282+j-(target_length-20)]=0

    "Determine whether GC is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='GC':
            gRNA_properties[301+j-(target_length-20)]=1
        else:
            gRNA_properties[301+j-(target_length-20)]=0

    "Determine whether CA is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='CA':
            gRNA_properties[320+j-(target_length-20)]=1
        else:
            gRNA_properties[320+j-(target_length-20)]=0

    "Determine whether CT is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='CT':
            gRNA_properties[339+j-(target_length-20)]=1
        else:
            gRNA_properties[339+j-(target_length-20)]=0

    "Determine whether CG is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='CG':
            gRNA_properties[358+j-(target_length-20)]=1
        else:
            gRNA_properties[358+j-(target_length-20)]=0

    "Determine whether CC is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='CC':
            gRNA_properties[376+j-(target_length-20)]=1
        else:
            gRNA_properties[376+j-(target_length-20)]=0

    return gRNA_properties




selected_target_strains=[]
selected_nontarget_strains=[]



#######################################################################################
############################# Graphical user interface ################################
#######################################################################################

from tkinter import *
from tkinter.filedialog import asksaveasfile
from tkinter import ttk
from tkinter.filedialog import askopenfilename

background_color="grey87"
button_color="light sky blue"

window = Tk() 
window.title("ssCRISPR")
window.geometry('1190x770')
window.config(background=background_color)

icon_name='Cas_icon3.ico'
icon_path = os.path.join(application_path, icon_name)
window.iconbitmap(icon_path)

lbl_title = Label(window,text = "ssCRISPR: Strain-specific CRISPR gRNA design software")

entry_email = Entry(window)
entry_email.insert(0, 'Enter email address. This is only used for NCBI searches.')

lbl2 = Label(window,text = "Select Cas protein and enter your PAM sequences as As, Ts, Cs, and Gs")
lbl = Label(window,text = "Select your strains")
lbl3 = Label(window,text = "Specify additional criteria")

btn_A = Button(window, text = "A", command=lambda *args: select_strains(A_strains_list))
btn_B = Button(window, text = "B", command=lambda *args: select_strains(B_strains_list))
btn_C = Button(window, text = "C", command=lambda *args: select_strains(C_strains_list))
btn_D = Button(window, text = "D", command=lambda *args: select_strains(D_strains_list))
btn_E = Button(window, text = "E", command=lambda *args: select_strains(E_strains_list))
btn_F = Button(window, text = "F", command=lambda *args: select_strains(F_strains_list))
btn_G = Button(window, text = "G", command=lambda *args: select_strains(G_strains_list))
btn_H = Button(window, text = "H", command=lambda *args: select_strains(H_strains_list))
btn_I = Button(window, text = "I", command=lambda *args: select_strains(I_strains_list))
btn_J = Button(window, text = "J", command=lambda *args: select_strains(J_strains_list))
btn_K = Button(window, text = "K", command=lambda *args: select_strains(K_strains_list))
btn_L = Button(window, text = "L", command=lambda *args: select_strains(L_strains_list))
btn_M = Button(window, text = "M", command=lambda *args: select_strains(M_strains_list))
btn_N = Button(window, text = "N", command=lambda *args: select_strains(N_strains_list))
btn_O = Button(window, text = "O", command=lambda *args: select_strains(O_strains_list))
btn_P = Button(window, text = "P", command=lambda *args: select_strains(P_strains_list))
btn_Q = Button(window, text = "Q", command=lambda *args: select_strains(Q_strains_list))
btn_R = Button(window, text = "R", command=lambda *args: select_strains(R_strains_list))
btn_S = Button(window, text = "S", command=lambda *args: select_strains(S_strains_list))
btn_T = Button(window, text = "T", command=lambda *args: select_strains(T_strains_list))
btn_U = Button(window, text = "U", command=lambda *args: select_strains(U_strains_list))
btn_V = Button(window, text = "V", command=lambda *args: select_strains(V_strains_list))
btn_W = Button(window, text = "W", command=lambda *args: select_strains(W_strains_list))
btn_X = Button(window, text = "X", command=lambda *args: select_strains(X_strains_list))
btn_Y = Button(window, text = "Y", command=lambda *args: select_strains(Y_strains_list))
btn_Z = Button(window, text = "Z", command=lambda *args: select_strains(Z_strains_list))
btn_A.config(font=("Arial", 10), width=2)
btn_B.config(font=("Arial", 10), width=2)
btn_C.config(font=("Arial", 10), width=2)
btn_D.config(font=("Arial", 10), width=2)
btn_E.config(font=("Arial", 10), width=2)
btn_F.config(font=("Arial", 10), width=2)
btn_G.config(font=("Arial", 10), width=2)
btn_H.config(font=("Arial", 10), width=2)
btn_I.config(font=("Arial", 10), width=2)
btn_J.config(font=("Arial", 10), width=2)
btn_K.config(font=("Arial", 10), width=2)
btn_L.config(font=("Arial", 10), width=2)
btn_M.config(font=("Arial", 10), width=2)
btn_N.config(font=("Arial", 10), width=2)
btn_O.config(font=("Arial", 10), width=2)
btn_P.config(font=("Arial", 10), width=2)
btn_Q.config(font=("Arial", 10), width=2)
btn_R.config(font=("Arial", 10), width=2)
btn_S.config(font=("Arial", 10), width=2)
btn_T.config(font=("Arial", 10), width=2)
btn_U.config(font=("Arial", 10), width=2)
btn_V.config(font=("Arial", 10), width=2)
btn_W.config(font=("Arial", 10), width=2)
btn_X.config(font=("Arial", 10), width=2)
btn_Y.config(font=("Arial", 10), width=2)
btn_Z.config(font=("Arial", 10), width=2)

lbl_nontargets = Label(window,text = "Non-target strains:")
lbl_all = Label(window,text = "Strains:")
lbl_targets = Label(window,text = "Target strains:")

entry_PAMs = Entry(window)
entry_PAMs.insert(0, 'Enter PAMs (ex: AGG, TGG, CGG, GGG)')

combobox_specificity = ttk.Combobox(window,values=['Nucleotides of specificity',1,2,3,4],state='readonly')
combobox_specificity.current(0)

combobox_nts = ttk.Combobox(window,values=['Target site length (bp)',20,21,22,23,24,25,26,27,28,29,30,31,32],state='readonly')
combobox_nts.current(0)

combobox_orientation = ttk.Combobox(window,values=['Orientation','5\'-target-PAM-3\'','5\'-PAM-target-3\''],state='readonly')
combobox_orientation.current(0)

listbox_all = Listbox(window, selectmode=EXTENDED)    
for strain in A_strains_list:
    listbox_all.insert(END,strain)
scrollbar_all = Scrollbar(window)    
  
listbox_targets = Listbox(window, selectmode=EXTENDED)
scrollbar_targets = Scrollbar(window)
listbox_nontargets = Listbox(window, selectmode=EXTENDED)
scrollbar_nontargets = Scrollbar(window)

def select_strains(letter_strains_list):
    global listbox_all
    listbox_all.delete(0,END)
    for strain in letter_strains_list:
        listbox_all.insert(END,strain) 


def select_target():
    index=list(listbox_all.curselection())
    index=index[::-1]
    for x in index:
        d=listbox_all.get(x)
        selected_target_strains.append(d)
        listbox_targets.insert(END,d)
        listbox_all.delete(x)
        
def add_custom_target():
    custom_target='User-defined target sequence'
    listbox_targets.insert(END,custom_target)
    selected_target_strains.append(custom_target)
    
def remove_target():
    index=list(listbox_targets.curselection())
    index=index[::-1]
    for x in index:
        d=listbox_targets.get(x)
        selected_target_strains.remove(d)
        listbox_targets.delete(x)
        
def select__nontarget():
    index=list(listbox_all.curselection())
    index=index[::-1]
    for x in index:
        d=listbox_all.get(x)
        selected_nontarget_strains.append(d)
        listbox_nontargets.insert(END,d)
        listbox_all.delete(x)
        
def add_custom_nontarget():
    custom_target='User-defined non-target sequence'
    listbox_nontargets.insert(END,custom_target)
    selected_nontarget_strains.append(custom_target)

def remove_nontarget():
    index=list(listbox_nontargets.curselection())
    index=index[::-1]
    for x in index:
        d=listbox_nontargets.get(x)
        selected_nontarget_strains.remove(d)
        listbox_nontargets.delete(x)
        
def reset_selections():
    global selected_target_strains, selected_nontarget_strains
    try:
        lbl_output.destroy()
    except:
        pass
    selected_target_strains=[]
    selected_nontarget_strains=[]
    
    global listbox_all, listbox_targets, listbox_nontargets
    listbox_all.delete(0,END) 
    for strain in A_strains_list:
        listbox_all.insert(END,strain)   
    listbox_targets.delete(0,END)
    listbox_nontargets.delete(0,END)
    
def download_results():
    files = [('Text Document', '*.txt')]
    file = asksaveasfile(filetypes = files, defaultextension = files) 
    
    file.write('PAM sequences:' + '\n')
    counter=1
    for PAM in PAM_seqs:
        file.write(str(counter)+'. ' + PAM + '\n')
        counter = counter+1
        
    if len(nonkill_strains_list)>0:
        file.write('\n' + 'Nucleotides of specificity:' + '\n')
        file.write(str(nucleotides_specificity) + '\n')
        
    file.write('\n' + 'gRNA target length:' + '\n')
    file.write(str(target_length) + '\n')
    
    file.write('\n' + 'gRNA orientation:' + '\n')
    file.write(str(PAM_orientation_output) + '\n')

    file.write('\n' + 'Target strains and sequences:' + '\n')
    counter=1
    for strain_name in kill_strains_list:
        file.write(str(counter)+'. ' + strain_name + '\n')
        counter = counter+1

    file.write('\n' + 'Nontarget strains and sequences:' + '\n')
    counter=1
    if len(nonkill_strains_list)==0:
        file.write('None' + '\n')
    else:
        for strain_name in nonkill_strains_list:
            file.write(str(counter)+'. ' + strain_name + '\n')
            counter = counter+1
    
    if Cpf1Check.get() == 1:
        file.write("\ngRNAs ranked by predicted Cpf1 efficiency:\n")
    elif Cas9Check.get() == 1:
        file.write("\ngRNAs ranked by predicted Cas9 efficiency:\n")
    else:
        file.write("\ngRNAs without efficiency ranking:\n")
    num=1
    for gRNA in all_gRNAs:
        file.write(str(num)+'. '+str(gRNA)+"\n")
        num=num+1
    file.write("\n")
    file.close()
    btn_download.destroy()
    
def Add_Cas9_PAM():
    if Cas9Check.get() == 1:
        Cpf1Check.set(0)
        OtherCheck.set(0)
        entry_PAMs.delete(0,END)
        entry_PAMs.insert(0, 'AGG, TGG, CGG, GGG')
        combobox_orientation.current(1)
    if Cas9Check.get() == 0:
        OtherCheck.set(1)
        entry_PAMs.delete(0,END)
        entry_PAMs.insert(0, 'Enter PAMs (ex: AGG, TGG, CGG, GGG)')
        combobox_orientation.current(0)
        
def Add_Cpf1_PAM():
    if Cpf1Check.get() == 1:
        Cas9Check.set(0)
        OtherCheck.set(0)
        entry_PAMs.delete(0,END)
        entry_PAMs.insert(0, 'TTTA, TTTC, TTTG')
        combobox_orientation.current(2)
    if Cpf1Check.get() == 0:
        OtherCheck.set(1)
        entry_PAMs.delete(0,END)
        entry_PAMs.insert(0, 'Enter PAMs (ex: AGG, TGG, CGG, GGG)')
        combobox_orientation.current(0)
        
def Add_other_PAM():
    if OtherCheck.get() == 1:
        Cas9Check.set(0)
        Cpf1Check.set(0)
        entry_PAMs.delete(0,END)
        entry_PAMs.insert(0, 'Enter PAMs (ex: AGG, TGG, CGG, GGG)')
        combobox_orientation.current(0)        
        
    
Cas9Check=IntVar()
Cas9Check_Box = Checkbutton(window, variable=Cas9Check, text='Cas9', font=("Arial", 12), bg = background_color, onvalue=1, offvalue=0, command=Add_Cas9_PAM)
Cpf1Check=IntVar()
Cpf1Check_Box = Checkbutton(window, variable=Cpf1Check, text='Cpf1', font=("Arial", 12), bg = background_color, onvalue=1, offvalue=0, command=Add_Cpf1_PAM)    
OtherCheck=IntVar(value=1)
OtherCheck_Box = Checkbutton(window, variable=OtherCheck, text='Other', font=("Arial", 12), bg = background_color, onvalue=1, offvalue=0, command=Add_other_PAM)

btn_select_target = Button(window, text = "Target strain >>", command = select_target)
btn_unselect_target = Button(window, text = "Remove", command = remove_target)
btn_select_nontarget = Button(window, text = "<< Don't target strain", command = select__nontarget)
btn_unselect_nontarget = Button(window, text = "Remove", command = remove_nontarget)
btn_custom_target = Button(window, text = "Add Fasta sequence", command = add_custom_target)
btn_custom_nontarget = Button(window, text = "Add Fasta sequence", command = add_custom_nontarget)

btn_generate_gRNAs = Button(window, text = "Determine gRNAs", command = lambda : determine_gRNAs(entry_PAMs.get(),selected_target_strains,selected_nontarget_strains,combobox_specificity.get(),combobox_nts.get(),combobox_orientation.get()))   
btn_reset = Button(window, text = "Reset selections", command = reset_selections)

lbl_title.config(font=("Arial", 18), background=background_color)
entry_email.config(font=("Arial", 12), background="white", width=50)
lbl2.config(font=("Arial", 14), background="LightCyan2", width=100)
entry_PAMs.config(font=("Arial", 12), background="white", width=50)
lbl3.config(font=("Arial", 14), background="LightCyan2", width=100)
combobox_specificity.config(font=("Arial", 12))
combobox_nts.config(font=("Arial", 12))
combobox_orientation.config(font=("Arial", 12))
lbl.config(font=("Arial", 14), background="LightCyan2", width=100)
lbl_nontargets.config(font=("Arial", 12), background=background_color)
lbl_all.config(font=("Arial", 12), background=background_color)
lbl_targets.config(font=("Arial", 12), background=background_color)
listbox_nontargets.config(font=("Arial", 10), width=51, height=12, yscrollcommand = scrollbar_nontargets.set) 
listbox_all.config(font=("Arial", 10), width=51, height=12, yscrollcommand = scrollbar_all.set)
listbox_targets.config(font=("Arial", 10), width=51, height=12, yscrollcommand = scrollbar_targets.set)
scrollbar_nontargets.config(command = listbox_nontargets.yview, width=20)
scrollbar_all.config(command = listbox_all.yview, width=20)
scrollbar_targets.config(command = listbox_targets.yview, width=20)
btn_select_target.config(font=("Arial", 12), width=20, background=button_color)
btn_unselect_target.config(font=("Arial", 12), width=20, background=button_color)
btn_select_nontarget.config(font=("Arial", 12), width=20, background=button_color)
btn_unselect_nontarget.config(font=("Arial", 12), width=20, background=button_color)
btn_generate_gRNAs.config(font=("Arial", 12), width=20, background="royal blue")
btn_reset.config(font=("Arial", 12), width=20, background="royal blue")
btn_custom_target.config(font=("Arial", 12), width=20, background=button_color)
btn_custom_nontarget.config(font=("Arial", 12), width=20, background=button_color)

lbl_title.grid(row=0,column=1,columnspan=28, pady=(8,12))
entry_email.grid(row=1, column=1, columnspan=28, pady=(4,4))
lbl2.grid(row=2,column=1,columnspan=28,pady=(4,8))
Cas9Check_Box.grid(row=3,column=4, columnspan=2)
Cpf1Check_Box.grid(row=3,column=6, columnspan=2)
OtherCheck_Box.grid(row=3,column=8, columnspan=2)
entry_PAMs.grid(row=3,column=5,columnspan=30)
lbl3.grid(row=4,column=1,columnspan=28,pady=(25,12))
combobox_specificity.grid(row=5,column=1,columnspan=8)
combobox_nts.grid(row=5,column=11,columnspan=8)
combobox_orientation.grid(row=5,column=21,columnspan=8)
lbl.grid(row=6,column=1,columnspan=28,pady=(25,0))
window.grid_rowconfigure(7, minsize=10)

btn_A.grid(row=8,column=2,pady=(10,10))
btn_B.grid(row=8,column=3)
btn_C.grid(row=8,column=4)
btn_D.grid(row=8,column=5)
btn_E.grid(row=8,column=6)
btn_F.grid(row=8,column=7)
btn_G.grid(row=8,column=8)
btn_H.grid(row=8,column=9)
btn_I.grid(row=8,column=10)
btn_J.grid(row=8,column=11)
btn_K.grid(row=8,column=12)
btn_L.grid(row=8,column=13)
btn_M.grid(row=8,column=14)
btn_N.grid(row=8,column=15)
btn_O.grid(row=8,column=16)
btn_P.grid(row=8,column=17)
btn_Q.grid(row=8,column=18)
btn_R.grid(row=8,column=19)
btn_S.grid(row=8,column=20)
btn_T.grid(row=8,column=21)
btn_U.grid(row=8,column=22)
btn_V.grid(row=8,column=23)
btn_W.grid(row=8,column=24)
btn_X.grid(row=8,column=25)
btn_Y.grid(row=8,column=26)
btn_Z.grid(row=8,column=27)

lbl_nontargets.grid(row=9,column=0,columnspan=10)
lbl_all.grid(row=9,column=10,columnspan=10)
lbl_targets.grid(row=9,column=20,columnspan=10)

listbox_nontargets.grid(row=10,column=0,columnspan=9)
scrollbar_nontargets.grid(row=10, column=9, sticky=W+N+S, padx=(0,20))
listbox_all.grid(row=10,column=10,columnspan=9)
scrollbar_all.grid(row=10, column=19, sticky=W+N+S, padx=(0,20))
listbox_targets.grid(row=10,column=20,columnspan=9)
scrollbar_targets.grid(row=10, column=29, sticky=W+N+S, padx=(0,20))

btn_custom_target.grid(row=11,column=20,pady=(4,2),columnspan=10)
btn_custom_nontarget.grid(row=11,column=0,pady=(4,2),columnspan=10)
btn_select_target.grid(row=11,column=10,pady=(4,2),columnspan=10)
btn_unselect_target.grid(row=12,column=20,pady=(4,2),columnspan=10)
btn_select_nontarget.grid(row=12,column=10,columnspan=10)
btn_unselect_nontarget.grid(row=12,column=0,pady=(4,2),columnspan=10)
window.grid_rowconfigure(13, minsize=20)
btn_generate_gRNAs.grid(row=14,column=10,columnspan=10)
btn_reset.grid(row=14,column=20,pady=(2,4),columnspan=10)

lbl_output = Label(window,text='',background=background_color)
window.grid_rowconfigure(15, minsize=10)
lbl_output.grid(row=16,column=10, columnspan=10)

 
window.mainloop()

