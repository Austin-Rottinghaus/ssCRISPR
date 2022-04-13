# -*- coding: utf-8 -*-
"""
Created on Tue Mar 22 13:33:03 2022

@author: Austin Rottinghaus
"""

from Bio.SeqUtils import nt_search
from Bio.Seq import Seq
from Bio.Alphabet import generic_dna, DNAAlphabet
import re
import time
import sys
import os
from openpyxl import load_workbook, cell, worksheet, workbook
import numpy
import pickle
import random
from Bio.SeqUtils import MeltingTemp as mt
from seqfold import *
from itertools import compress 
import math
import numpy as np

import matplotlib.pyplot as plt
from sklearn import datasets, ensemble
from sklearn.inspection import permutation_importance
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split


wb=load_workbook('Cpf1 Dataset.xlsx')
ws=wb['Data']
prediction_model = pickle.load(open('finalized_model.sav','rb'))
target_length=20




def property_identifier_5primePAM(gRNA):
    gRNA_properties=[0]*396

    "Check num of Ts in -4 to -1 from PAM"
    numTs_14=0
    for j in [0,1,2,3]:
        if gRNA[j]=='T':
            numTs_14=numTs_14+1
    gRNA_properties[0]=numTs_14

    "Determine GC content"
    numGCs=0
    for j in range(0,target_length):
        if gRNA[j]=='G' or gRNA[j]=='C':
            numGCs=numGCs+1
    gRNA_properties[1]=numGCs/20

    "Determine A content"
    numAs=0
    for j in range(0,target_length):
        if gRNA[j]=='A':
            numAs=numAs+1
    gRNA_properties[2]=numAs/20

    "Determine T content"
    numTs=0
    for j in range(0,target_length):
        if gRNA[j]=='T':
            numTs=numTs+1
    gRNA_properties[3]=numTs/20

    "Determine G content"
    numGs=0
    for j in range(0,target_length):
        if gRNA[j]=='G':
            numGs=numGs+1
    gRNA_properties[4]=numGs/20

    "Determine C content"
    numCs=0
    for j in range(0,target_length):
        if gRNA[j]=='C':
            numCs=numCs+1
    gRNA_properties[5]=numCs/20

    "Determine free energy of full RNA sequence"
    gRNA_RNA=str(gRNA).replace('T','U')
    try:
        dg(gRNA_RNA, temp = 37.0)
    except:
        print(gRNA)
        print(len(gRNA))
        print(gRNA_RNA)
        print(len(gRNA_RNA))
        print([numAs,numTs,numGs,numCs])

    if dg(gRNA_RNA, temp = 37.0) > -100:
        if dg(gRNA_RNA, temp = 37.0) < 4:
            gRNA_properties[6]=dg(gRNA_RNA, temp = 37.0)
        else:
            gRNA_properties[6]=4
    else:
        gRNA_properties[6]=4


    "Determine free energy of the 12bp of RNA closest to the PAM"
    gRNA_RNA_short=gRNA_RNA[0:12]
    if dg(gRNA_RNA_short, temp = 37.0) > -30:
        if dg(gRNA_RNA_short, temp = 37.0) < 4:
            gRNA_properties[7]=dg(gRNA_RNA_short, temp = 37.0)
        elif str(abs(dg(gRNA_RNA_short, temp = 37.0))) == 'inf':
            gRNA_properties[7]=4
        else:
            gRNA_properties[7]=4
    else:
        gRNA_properties[7]=4

    "Calculate melting temp for full gRNA"
    gRNA_properties[8]=mt.Tm_NN(Seq(str(gRNA)).complement(), c_seq=Seq(gRNA_RNA))
    "Calculate melting temp for 5bp closest to PAM"
    gRNA_properties[9]=mt.Tm_NN(Seq(str(gRNA)).complement()[0:5], c_seq=Seq(gRNA_RNA)[0:5])
    "Calculate melting temp for middle 8bp"
    gRNA_properties[10]=mt.Tm_NN(Seq(str(gRNA)).complement()[5:13], c_seq=Seq(gRNA_RNA)[5:13])
    "Calculate melting temp for remaining bp"
    gRNA_properties[11]=mt.Tm_NN(Seq(str(gRNA)).complement()[13:target_length], c_seq=Seq(gRNA_RNA)[13:target_length])

    "Determine whether an A is in each position of 20bp near PAM"
    for j in range(19,-1,-1):
        if gRNA[j]=='A':
            gRNA_properties[12+j-(target_length-20)]=1
        else:
            gRNA_properties[12+j-(target_length-20)]=0

    "Determine whether a T is in each position"
    for j in range(19,-1,-1):
        if gRNA[j]=='T':
            gRNA_properties[32+j-(target_length-20)]=1
        else:
            gRNA_properties[32+j-(target_length-20)]=0

    "Determine whether a G is in each position"
    for j in range(19,-1,-1):
        if gRNA[j]=='G':
            gRNA_properties[52+j-(target_length-20)]=1
        else:
            gRNA_properties[52+j-(target_length-20)]=0

    "Determine whether a C is in each position"
    for j in range(19,-1,-1):
        if gRNA[j]=='C':
            gRNA_properties[72+j-(target_length-20)]=1
        else:
            gRNA_properties[72+j-(target_length-20)]=0

    "Determine whether AA is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='AA':
            gRNA_properties[92+j-(target_length-20)]=1
        else:
            gRNA_properties[92+j-(target_length-20)]=0

    "Determine whether AT is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='AT':
            gRNA_properties[111+j-(target_length-20)]=1
        else:
            gRNA_properties[111+j-(target_length-20)]=0

    "Determine whether AG is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='AG':
            gRNA_properties[130+j-(target_length-20)]=1
        else:
            gRNA_properties[130+j-(target_length-20)]=0

    "Determine whether AC is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='AC':
            gRNA_properties[149+j-(target_length-20)]=1
        else:
            gRNA_properties[149+j-(target_length-20)]=0

    "Determine whether TA is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='TA':
            gRNA_properties[168+j-(target_length-20)]=1
        else:
            gRNA_properties[168+j-(target_length-20)]=0

    "Determine whether TT is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='TT':
            gRNA_properties[187+j-(target_length-20)]=1
        else:
            gRNA_properties[187+j-(target_length-20)]=0

    "Determine whether TG is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='TG':
            gRNA_properties[206+j-(target_length-20)]=1
        else:
            gRNA_properties[206+j-(target_length-20)]=0

    "Determine whether TC is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='TC':
            gRNA_properties[225+j-(target_length-20)]=1
        else:
            gRNA_properties[225+j-(target_length-20)]=0

    "Determine whether GA is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='GA':
            gRNA_properties[244+j-(target_length-20)]=1
        else:
            gRNA_properties[244+j-(target_length-20)]=0

    "Determine whether GT is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='GT':
            gRNA_properties[263+j-(target_length-20)]=1
        else:
            gRNA_properties[263+j-(target_length-20)]=0

    "Determine whether GG is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='GG':
            gRNA_properties[282+j-(target_length-20)]=1
        else:
            gRNA_properties[282+j-(target_length-20)]=0

    "Determine whether GC is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='GC':
            gRNA_properties[301+j-(target_length-20)]=1
        else:
            gRNA_properties[301+j-(target_length-20)]=0

    "Determine whether CA is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='CA':
            gRNA_properties[320+j-(target_length-20)]=1
        else:
            gRNA_properties[320+j-(target_length-20)]=0

    "Determine whether CT is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='CT':
            gRNA_properties[339+j-(target_length-20)]=1
        else:
            gRNA_properties[339+j-(target_length-20)]=0

    "Determine whether CG is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='CG':
            gRNA_properties[358+j-(target_length-20)]=1
        else:
            gRNA_properties[358+j-(target_length-20)]=0

    "Determine whether CC is in each 2x position"
    for j in range(19,0,-1):
        if gRNA[j-1:j+1]=='CC':
            gRNA_properties[376+j-(target_length-20)]=1
        else:
            gRNA_properties[376+j-(target_length-20)]=0

    return gRNA_properties

sequences=[]

for i in range(2,15002):
    sequences.append(ws.cell(i,1).value)
print(len(sequences))
    
properties=list(map(property_identifier_5primePAM,sequences))
efficiencies=[]
for i in range(2,15002):
    mycell=ws.cell(row=i, column=2)   
    efficiencies.append(mycell.value)

wb.save('Cpf1 Dataset.xlsx')
wb.close()


# Split data into 90% train and 10% test    
sgRNA_properties_train, sgRNA_properties_test, log2scores_train, log2scores_test = train_test_split(
    properties, efficiencies, test_size=0.1, random_state=13)



# Set parameters for the gradient boosting regression
params = {'n_estimators': 400, #
          'max_depth': 7, #
          'min_samples_split': 75, #
          'learning_rate': .03, #
          'loss': 'ls'}

# Perform the regression
reg = ensemble.GradientBoostingRegressor(**params)
reg.fit(sgRNA_properties_train, log2scores_train)
print('400,7,80')
mse = mean_squared_error(log2scores_test, reg.predict(sgRNA_properties_test))
print("The mean squared error (MSE) on test set: {:.4f}".format(mse))



pickle.dump(reg, open('finalized_model_cpf1.sav', 'wb'))

#all_predictions = reg.predict(sgRNA_properties)
#for i in range(2,56337):
#    ws_predict.cell(i,3).value=all_predictions[i-2]


# Plot the deviance throughout the successive boosting iterations
test_score = np.zeros((params['n_estimators'],), dtype=np.float64)
for i, log2scores_pred in enumerate(reg.staged_predict(sgRNA_properties_test)):
    test_score[i] = reg.loss_(log2scores_test, log2scores_pred)

fig = plt.figure(figsize=(6, 6))
plt.subplot(1, 1, 1)
plt.title('Deviance')
plt.plot(np.arange(params['n_estimators']) + 1, reg.train_score_, 'b-',
         label='Training Set Deviance')
plt.plot(np.arange(params['n_estimators']) + 1, test_score, 'r-',
         label='Test Set Deviance')
plt.legend(loc='upper right')
plt.xlabel('Boosting Iterations')
plt.ylabel('Deviance')
fig.tight_layout()
plt.show()
